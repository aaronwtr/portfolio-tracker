import os
import pandas as pd
import openpyxl as pyxl


class DegiroUpdateCSV:
    def __init__(self):
        with open("len_products.txt", "r") as f:
            self.products = int(f.read())
        f.close()

        self.portfolio_path = os.getenv("PORTFOLIO_CSV")
        self.header_value = 15  # Start of stock portfolio rows

    def get_excel_stocks(self):
        stocks_portfolio = pd.read_excel(self.portfolio_path, sheet_name='Stocks portfolio',
                                         header=self.header_value - 2, usecols="A:J", nrows=self.products)

        excel_stocks = list(stocks_portfolio["Code"])
        stocks_value_old = list(stocks_portfolio["Huidige waarde"])
        dict_old_value = dict(zip(excel_stocks, stocks_value_old))
        return excel_stocks, dict_old_value

    def update_stocks(self, excel_stocks, dict_old_value, DGF, save=False):
        wb = pyxl.load_workbook(filename=self.portfolio_path)
        ws = wb.worksheets[0]

        products = DGF.fetch_portfolio()
        # print(products)
        # ws['E6'].value = products['EUR']

        if not os.path.isfile("len_products.txt"):
            with open("len_products.txt", "w") as f:
                f.write(str(len(products)))
            f.close()

        curr_row = 15
        excel_stock_loc = {}
        for i in range(len(excel_stocks)):
            excel_stock_loc[curr_row] = excel_stocks[i]
            curr_row += 1

        output_summary = []
        for key, value in excel_stock_loc.items():
            price_new = round(products[value], 2)
            ws['J{}'.format(key)].value = price_new

            price_old = round(dict_old_value[value], 2)

            percentage_change = round(((price_new - price_old)/price_old)*100, 2)
            if percentage_change > 0:
                percentage_change = '+{}'.format(percentage_change)

            output_summary.append('{}: {} --> {} ({}%)'.format(value, price_old, price_new, percentage_change))
        #
        if save:
            wb.save(self.portfolio_path)    # In order to make changes have effect, put save = True
            print('Success! The new stock values were saved to {}\n\nThe following changes were made:\n'.format(
                self.portfolio_path))
        else:
            print('Success! The results were not saved. If you want to save the results, add save=True to update_stocks().'
                  '\n\nThe following changes were made:'.format(self.portfolio_path))

        for summary in output_summary:
            print(summary)

        print('\n')


class CryptoUpdateCSV:
    def __init__(self, portfolio_path=None):
        self.portfolio_path = portfolio_path if portfolio_path else os.getenv("PORTFOLIO_CSV")

    def get_portfolio_coins(self):
        """Read the crypto portfolio from Excel file and return coins with their current values"""
        try:
            # Read the Excel file with pandas
            coins_portfolio = pd.read_excel(
                self.portfolio_path,
                sheet_name='Crypto portfolio',
                header=9,  # Start reading from the 10th row (0-indexed)
                usecols="A:E",
                nrows=18
            )

            # Extract coin codes and values
            excel_coins = list(coins_portfolio["Code"])
            coins_value_old = list(coins_portfolio["Huidige waarde"])

            # Create dictionary of coin codes and values
            dict_old_coin_value = dict(zip(excel_coins, coins_value_old))

            # Filter out coins with zero value
            dict_old_coin_value = {k: v for k, v in dict_old_coin_value.items() if
                                   pd.notna(v) and (isinstance(v, (int, float)) and v != 0 or
                                                    isinstance(v, str) and self._convert_to_float(v) != 0)}

            return dict_old_coin_value
        except Exception as e:
            print(f"Error reading Excel file: {str(e)}")
            return {}

    def _convert_to_float(self, value):
        """Convert currency string to float"""
        if pd.isna(value):
            return 0.0
        if isinstance(value, (int, float)):
            return float(value)
        # Remove € symbol, spaces, and replace comma with dot
        value = str(value).replace('€', '').replace(' ', '').replace(',', '.')
        try:
            return float(value)
        except ValueError:
            return 0.0

    def update_portfolio(self, bitv_port, bin_port=None, save=False):
        """Update the Excel file with values from Bitvavo and Binance"""
        try:
            # Get original values for comparison
            dict_old_value = self.get_portfolio_coins()

            # Load the workbook
            wb = pyxl.load_workbook(filename=self.portfolio_path)
            ws = wb.worksheets[1]  # Assuming 'Crypto portfolio' is the second sheet (0-indexed)

            # Combine portfolio data from different exchanges
            crypto_portfolio = {}
            for d in (bitv_port, bin_port or {}):
                for k, v in d.items():
                    if v != 0.0:
                        if k in crypto_portfolio:
                            crypto_portfolio[k] += v
                        else:
                            crypto_portfolio[k] = v

            # Find coin locations in the Excel sheet
            coin_rows = {}
            for row in range(11, 30):  # Adjust range as needed
                cell_value = ws.cell(row=row, column=2).value
                if cell_value:
                    coin_rows[cell_value] = row

            # Update coin values
            output_summary = []
            total_value = 0

            for coin_code, value in crypto_portfolio.items():
                if coin_code in coin_rows:
                    row = coin_rows[coin_code]

                    # Format value as currency
                    price_new = round(value, 2)
                    total_value += price_new

                    # Update cell with new value
                    ws.cell(row=row, column=5).value = price_new

                    # Calculate percentage change for reporting
                    if coin_code in dict_old_value:
                        price_old = self._convert_to_float(dict_old_value[coin_code])
                        if price_old > 0:
                            percentage_change = round(((price_new - price_old) / price_old) * 100, 2)
                            change_str = f"+{percentage_change}%" if percentage_change > 0 else f"{percentage_change}%"
                            output_summary.append(f"{coin_code}: {price_old:.2f} --> {price_new:.2f} ({change_str})")
                        else:
                            output_summary.append(f"{coin_code}: 0 --> {price_new:.2f} (new value)")
                    else:
                        output_summary.append(f"{coin_code}: NEW --> {price_new:.2f} (new asset)")

            # Update the total
            total_row = None
            for row in range(11, 40):
                if ws.cell(row=row, column=1).value == "TOTAL":
                    total_row = row
                    break

            if total_row:
                ws.cell(row=total_row, column=5).value = total_value

            # Save if requested
            if save:
                wb.save(self.portfolio_path)
                print(f'Success! The new coin values were saved to {self.portfolio_path}')
            else:
                print(
                    'Success! The results were not saved. If you want to save the results, add save=True to update_portfolio().')

            print('\nThe following changes were made:')
            for summary in output_summary:
                print(summary)

            print(f"\nTotal portfolio value updated to: € {total_value:.2f}")

            return True

        except Exception as e:
            print(f"Error updating Excel file: {str(e)}")
            return False
